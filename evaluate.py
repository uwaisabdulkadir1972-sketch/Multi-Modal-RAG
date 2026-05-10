"""
evaluate.py — Local MM-RAG Evaluation Script
=============================================
Runs the full golden QA evaluation against your local mmrag_store index.
Uses llama3.1:8b via Ollama as the judge (no API key needed).
Exports results to mmrag_evaluation.xlsx in the current directory.

Prerequisites:
  1. ollama pull llama3.1:8b
  2. ollama serve  (in a separate terminal)
  3. Run: python evaluate.py  (with venv activated)
"""

import os, re, time, datetime, json, sys
from pathlib import Path
from collections import defaultdict

# ── Make sure backend is importable ───────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))
from backend import MMRAGBackend

# ─────────────────────────────────────────────────────────────────────────
# GOLDEN QA
# ─────────────────────────────────────────────────────────────────────────

golden_qa = [

    # ── Tier 1: Direct Text ───────────────────────────────────────────────
    {
    "tier":     "1_direct_text",
    "question": "How many new patent applications did Infineon file in fiscal year 2024?",
    "answer":   "Approximately 1,900 new patent applications.",
    "pages":    ["fy2025-q3-web-en_page_025.png"],
    "facts":    ["1,900"],
    },
    {
        "tier":     "1_direct_text",
        "question": "What is Infineon's most recent quarterly revenue?",
        "answer":   "EUR 3,704 million in Q3 FY25.",
        "pages":    ["fy2025-q3-web-en_page_010.png"],
        "facts":    ["3,704"],
    },
    {
        "tier":     "1_direct_text",
        "question": "How many patents does Infineon hold in its portfolio?",
        "answer":   "29,900 patents and patent applications.",
        "pages":    ["fy2025-q3-web-en_page_025.png"],
        "facts":    ["29,900"],
    },
    {
        "tier":     "1_direct_text",
        "question": "What are the core applications served by the Green Industrial Power segment?",
        "answer":   (
            "Energy generation, energy storage, energy transmission, home appliances, "
            "industrial drives, industrial power supplies, industrial robotics, "
            "industrial vehicles, and traction."
        ),
        "pages":    ["fy2025-q3-web-en_page_017.png"],
        "facts":    ["energy generation", "energy storage", "industrial drives", "traction"],
    },

    # ── Tier 2: Visual Read ───────────────────────────────────────────────
    {
        "tier":     "2_visual_read",
        "question": "What are Infineon's market share percentages in Automotive semiconductors and the Microcontroller market?",
        "answer":   "13.5% in Automotive semiconductors and 21.4% in Microcontrollers.",
        "pages":    ["fy2025-q3-web-en_page_008.png"],
        "facts":    ["13.5", "21.4"],
    },
    {
        "tier":     "2_visual_read",
        "question": "Which region contributes the largest share of Infineon's revenue, and what is that percentage?",
        "answer":   "Mainland China and Hong Kong at 27%.",
        "pages":    ["fy2025-q3-web-en_page_012.png"],
        "facts":    ["27", "Mainland China"],
    },
    {
        "tier":     "2_visual_read",
        "question": "Has Infineon's total revenue recovered to its year-ago level?",
        "answer":   "Yes, essentially flat. Q3 FY24 was EUR 3,702 million and Q3 FY25 is EUR 3,704 million.",
        "pages":    ["fy2025-q3-web-en_page_010.png"],
        "facts":    ["3,702", "3,704"],
    },
    {
        "tier":     "2_visual_read",
        "question": "Did Infineon's revenue grow or decline from FY23 to FY24?",
        "answer":   "Declined, from EUR 16,309 million in FY23 to EUR 14,955 million in FY24.",
        "pages":    ["fy2025-q3-web-en_page_004.png"],
        "facts":    ["16,309", "14,955"],
    },

    # ── Tier 3: Multi-Slide ───────────────────────────────────────────────
    {
        "tier":     "3_multi_slide",
        "question": "Which of Infineon's four segments has the highest and lowest segment result margin in Q3 FY25?",
        "answer":   "Highest is Automotive at 19.8%. Lowest is Connected Secure Systems at 11.2%.",
        "pages":    ["fy2025-q3-web-en_page_015.png",
                     "fy2025-q3-web-en_page_017.png",
                     "fy2025-q3-web-en_page_019.png",
                     "fy2025-q3-web-en_page_021.png"],
        "facts":    ["19.8", "11.2", "Automotive", "Connected Secure Systems"],
    },
    {
        "tier":     "3_multi_slide",
        "question": "Which segment improved its margin the most from Q3 FY24 to Q3 FY25?",
        "answer":   "Power & Sensor Systems, from 11.7% to 18.8%, an improvement of 7.1 percentage points.",
        "pages":    ["fy2025-q3-web-en_page_019.png"],
        "facts":    ["Power & Sensor Systems", "11.7", "18.8"],
    },
    {
        "tier":     "3_multi_slide",
        "question": "Infineon claims to be number one in automotive semiconductors. Who is their closest competitor and what is the gap?",
        "answer":   "NXP is the closest competitor at 10.5%, versus Infineon's 13.5%, a gap of 3 percentage points.",
        "pages":    ["fy2025-q3-web-en_page_003.png",
                     "fy2025-q3-web-en_page_008.png"],
        "facts":    ["13.5", "NXP", "10.5"],
    },
    {
        "tier":     "3_multi_slide",
        "question": "What share of Infineon's FY24 revenue comes from the growth areas the company highlights?",
        "answer":   (
            "40% in total: E-mobility 16%, Renewables 7%, Software-defined vehicle 7%, "
            "IoT 5%, and AI/Data center 5%."
        ),
        "pages":    ["fy2025-q3-web-en_page_004.png",
                     "fy2025-q3-web-en_page_013.png"],
        "facts":    ["40", "E-mobility", "Renewables", "IoT", "AI"],
    },

    # ── Tier 4: Stress Tests ──────────────────────────────────────────────
    {
        "tier":     "4_stress",
        "question": "What is the forecast revenue range for the global semiconductor market in 2026?",
        "answer":   "USD 761 billion to USD 839 billion.",
        "pages":    ["fy2025-q3-web-en_page_007.png"],
        "facts":    ["761", "839"],
    },
    {
        "tier":     "4_stress",
        "question": "What is Infineon's CO2 burden-to-savings ratio and what are the exact tonnage figures?",
        "answer":   "Ratio of 1:45. Burden of 2.9 million tons, savings of 130.3 million tons CO2 equivalents.",
        "pages":    ["fy2025-q3-web-en_page_029.png"],
        "facts":    ["1:45", "2.9", "130.3"],
    },

]

# ─────────────────────────────────────────────────────────────────────────
# JUDGE — llama-3.3-70b via Groq (free, strong, independent from generator)
# ─────────────────────────────────────────────────────────────────────────
import requests as _requests

GROQ_JUDGE_KEY   = ""  # ← paste your Groq key here
GROQ_JUDGE_MODEL = "llama-3.3-70b-versatile"
GROQ_JUDGE_URL   = "https://api.groq.com/openai/v1/chat/completions"

_STOPWORDS = {
    "a","an","the","and","or","of","in","to","for","is","are","was","were",
    "it","its","that","this","with","from","by","at","on","as","be","been",
    "has","have","had","not","but","will","would","could","should","may",
    "might","do","did","does","what","how","when","where","which","who","also"
}

def _keywords(text):
    return set(w for w in re.findall(r"\w+", text.lower())
               if w not in _STOPWORDS and len(w) > 2)

_JUDGE_PROMPT = (
    "You are a fair evaluator of a question-answering system.\n\n"
    "Question        : {question}\n"
    "Reference answer: {reference}\n"
    "System answer   : {prediction}\n\n"
    "Evaluate the System answer. Be lenient about paraphrasing.\n"
    "  GOOD = Correct and covers the main points. Paraphrases are fine.\n"
    "  OK   = Mostly right but missing some details, or a minor inaccuracy.\n"
    "  BAD  = Wrong, off-topic, hallucinated, or repeats without addressing the question.\n\n"
    "Reply with ONE word only: GOOD, OK, or BAD."
)

_QUALITY_SCORES = {"GOOD": 1.0, "OK": 0.5, "BAD": 0.0}

def _check_ollama_judge():
    if GROQ_JUDGE_KEY == "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" or not GROQ_JUDGE_KEY:
        print("⚠️  Groq judge key not set — falling back to word-overlap scoring")
        print("   Set GROQ_JUDGE_KEY in evaluate.py")
        return False
    try:
        r = _requests.get(
            "https://api.groq.com/openai/v1/models",
            headers={"Authorization": f"Bearer {GROQ_JUDGE_KEY}"},
            timeout=5
        )
        r.raise_for_status()
        print(f"✅ Judge ready: {GROQ_JUDGE_MODEL} via Groq")
        return True
    except Exception as e:
        print(f"⚠️  Groq judge unavailable ({e}) — falling back to word-overlap scoring")
        return False

def _word_overlap_quality(reference, prediction):
    ref_words  = _keywords(reference)
    pred_words = _keywords(prediction)
    if not ref_words: return 0.5, "OK*"
    if not pred_words: return 0.0, "BAD*"
    len_ratio = len(prediction) / max(len(reference), 1)
    recall    = len(ref_words & pred_words) / len(ref_words)
    if len_ratio > 6: recall *= 0.4
    if recall >= 0.55: return 1.0, "GOOD*"
    elif recall >= 0.28: return 0.5, "OK*"
    else: return 0.0, "BAD*"

_JUDGE_AVAILABLE = _check_ollama_judge()

def answer_quality(question, reference, prediction):
    if not _JUDGE_AVAILABLE:
        return _word_overlap_quality(reference, prediction)
    prompt = _JUDGE_PROMPT.format(
        question=question, reference=reference, prediction=prediction)
    for attempt in range(3):
        try:
            resp = _requests.post(
                GROQ_JUDGE_URL,
                headers={
                    "Authorization": f"Bearer {GROQ_JUDGE_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model":       GROQ_JUDGE_MODEL,
                    "messages":    [{"role": "user", "content": prompt}],
                    "max_tokens":  10,
                    "temperature": 0,
                },
                timeout=30,
            )
            resp.raise_for_status()
            raw = resp.json()["choices"][0]["message"]["content"].strip().upper()
            for key in ["GOOD", "OK", "BAD"]:
                if key in raw:
                    return _QUALITY_SCORES[key], key
            return 0.5, "OK"
        except Exception as e:
            if attempt < 2:
                time.sleep(2 ** attempt)
            else:
                print(f"  ⚠️  Judge error — using word-overlap: {e}")
    return _word_overlap_quality(reference, prediction)

# ─────────────────────────────────────────────────────────────────────────
# METRICS
# ─────────────────────────────────────────────────────────────────────────
def retrieval_hit_and_rank(retrieved_pages, gold_pages):
    if not gold_pages:
        return None, None
    def stem(p):
        return os.path.splitext(os.path.basename(str(p)).lower())[0]
    gold_stems = {stem(p) for p in gold_pages}
    for i, rp in enumerate(retrieved_pages, 1):
        if stem(rp) in gold_stems:
            return 1.0, i
    return 0.0, None

def retrieval_recall_and_rank(retrieved_pages, gold_pages):
    """Fraction of gold pages retrieved (recall), plus rank of first hit.
    Used for Tier 3 multi-slide questions where partial retrieval should
    be rewarded rather than scoring binary 0 or 1.
    """
    if not gold_pages:
        return None, None
    def stem(p):
        return os.path.splitext(os.path.basename(str(p)).lower())[0]
    gold_stems = {stem(p) for p in gold_pages}
    found      = 0
    first_rank = None
    for i, rp in enumerate(retrieved_pages, 1):
        if stem(rp) in gold_stems:
            found += 1
            if first_rank is None:
                first_rank = i
    recall = round(found / len(gold_pages), 3)
    return recall, first_rank

def _fact_matches(fact, answer_lower):
    if fact.lower() in answer_lower:
        return True
    words = [w for w in re.findall(r"\w+", fact.lower())
             if w not in _STOPWORDS and len(w) > 2]
    if not words:
        return fact.lower() in answer_lower
    found = sum(1 for w in words if w in answer_lower)
    return (found / len(words)) >= 0.75

def fact_coverage(facts, answer):
    if not facts:
        return None, []
    answer_lower = answer.lower()
    missing = [f for f in facts if not _fact_matches(f, answer_lower)]
    return round(1.0 - len(missing) / len(facts), 3), missing

def overall_score(hit, coverage, quality):
    scores = [s for s in [hit, coverage, quality] if s is not None]
    return round(sum(scores) / len(scores), 3) if scores else None

# ─────────────────────────────────────────────────────────────────────────
# RUN EVALUATION
# ─────────────────────────────────────────────────────────────────────────
def run_evaluation(backend):
    TIER_DISPLAY = {
        "1_direct_text": "T1 Direct",
        "2_visual_read": "T2 Visual",
        "3_multi_slide": "T3 Multi",
        "4_stress":      "T4 Stress",
    }

    results = []
    print()
    print(f"{'─'*92}")
    print(f"  {'#':>3}  {'Tier':<11}  {'Question':<34}  {'Hit':>5}  {'Rank':>4}  {'Cov':>6}  {'Quality':>7}  {'Score':>6}")
    print(f"{'─'*92}")

    for idx, qa in enumerate(golden_qa, 1):
        question  = qa["question"]
        reference = qa["answer"]
        pages     = qa.get("pages", [])
        facts     = qa.get("facts", [])
        tier      = qa.get("tier", "unknown")

        t0         = time.time()
        result     = backend.ask(question, top_k=5)
        latency_s  = round(time.time() - t0, 1)
        prediction = result["answer"]
        retrieved  = [s["filename"] for s in result["sources"]]

        if tier == "3_multi_slide" and len(pages) > 1:
            hit, rank = retrieval_recall_and_rank(retrieved, pages)
        else:
            hit, rank = retrieval_hit_and_rank(retrieved, pages)
        cov, missing  = fact_coverage(facts, prediction)
        qual, verdict = answer_quality(question, reference, prediction)
        score         = overall_score(hit, cov, qual)

        results.append({
            "idx":             idx,
            "tier":            tier,
            "question":        question,
            "reference":       reference,
            "prediction":      prediction,
            "retrieved":       retrieved,
            "gold_pages":      pages,
            "hit":             hit,
            "retrieval_rank":  rank,
            "fact_coverage":   cov,
            "missing_facts":   missing,
            "quality_score":   qual,
            "quality_verdict": verdict,
            "overall":         score,
            "latency_s":       latency_s,
        })

        tier_lbl = TIER_DISPLAY.get(tier, tier[:9])
        hit_str  = ("✓" if hit == 1.0 else "✗") if hit is not None else "--"
        rank_str = f"#{rank}" if rank else "--"
        cov_str  = f"{cov:.0%}" if cov is not None else "--"
        ov_str   = f"{score:.2f}" if score is not None else "--"
        note     = "  [excl. avg]" if tier == "4_stress" else ""
        print(f"  {idx:>3}.  {tier_lbl:<11}  {question[:34]:<34}  "
              f"{hit_str:>5}  {rank_str:>4}  {cov_str:>6}  {verdict:>7}  {ov_str:>6}{note}")

    print(f"{'─'*92}")

    # Summary
    non_t4 = [r for r in results if r.get("tier") != "4_stress"]
    hits   = [r["hit"]           for r in non_t4 if r["hit"]           is not None]
    covs   = [r["fact_coverage"] for r in non_t4 if r["fact_coverage"] is not None]
    quals  = [r["quality_score"] for r in non_t4 if r["quality_score"] is not None]
    overs  = [r["overall"]       for r in non_t4 if r["overall"]       is not None]
    lats   = [r["latency_s"]     for r in results]
    rnks   = [r["retrieval_rank"] for r in non_t4
              if r.get("retrieval_rank") is not None and r.get("hit") == 1.0]
    mrr    = sum(1.0/x for x in rnks) / len(non_t4) if rnks else None

    print()
    print("OVERALL SUMMARY  (Tier 1-3 only; Tier 4 excluded)")
    print(f"{'─'*55}")
    if hits:  print(f"  Hit Rate   : {sum(hits)/len(hits):.0%}  ({int(sum(hits))}/{len(hits)})")
    if mrr:   print(f"  MRR        : {mrr:.3f}")
    if covs:  print(f"  Fact Cov.  : {sum(covs)/len(covs):.0%}")
    if quals: print(f"  Quality    : {sum(quals)/len(quals):.2f}")
    if overs: print(f"  Overall    : {sum(overs)/len(overs):.2f}")
    if lats:  print(f"  Latency    : {sum(lats)/len(lats):.1f}s per question")

    return results

# ─────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────
def export_excel(results, output_path="mmrag_evaluation.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C = {
        "navy":"1F3864","navy2":"2E4A7A","steel":"4472C4",
        "lblue":"D6E4F0","white":"FFFFFF","offwh":"F8F9FA",
        "gbg":"E2EFDA","gfg":"2E7D32",
        "abg":"FFF2CC","afg":"E65100",
        "rbg":"FCE4D6","rfg":"C00000",
        "t1bg":"E8F5E9","t1fg":"1B5E20",
        "t2bg":"FFFDE7","t2fg":"E65100",
        "t3bg":"FFF3E0","t3fg":"BF360C",
        "t4bg":"FFEBEE","t4fg":"B71C1C",
        "grayhd":"D9D9D9","graybg":"F5F5F5",
    }
    TIER_META = {
        "1_direct_text": {"label":"Tier 1 -- Direct Text",  "note":">= 90%",           "bg":"t1bg","fg":"t1fg"},
        "2_visual_read": {"label":"Tier 2 -- Visual Read",  "note":"target 60-80%",     "bg":"t2bg","fg":"t2fg"},
        "3_multi_slide": {"label":"Tier 3 -- Multi-Slide",  "note":"target 40-65%",     "bg":"t3bg","fg":"t3fg"},
        "4_stress":      {"label":"Tier 4 -- Stress Tests", "note":"excluded from avg", "bg":"t4bg","fg":"t4fg"},
    }
    def F(h):  return PatternFill("solid", fgColor=C.get(h, h))
    def FT(bold=False, color="222222", size=10, italic=False):
        return Font(bold=bold, color=C.get(color,color), size=size, italic=italic, name="Segoe UI")
    def AL(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def BD():
        s = Side(style="thin", color="BFBFBF")
        return Border(bottom=s)
    def Pct(v): return f"{v:.0%}" if v is not None else "--"
    def CW(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    non_t4 = [r for r in results if r.get("tier") != "4_stress"]

    def _smry(res):
        nt4  = [r for r in res if r.get("tier") != "4_stress"]
        hits  = [r["hit"]           for r in nt4 if r.get("hit")           is not None]
        covs  = [r["fact_coverage"] for r in nt4 if r.get("fact_coverage") is not None]
        quals = [r["quality_score"] for r in nt4 if r.get("quality_score") is not None]
        overs = [r["overall"]       for r in nt4 if r.get("overall")       is not None]
        lats  = [r.get("latency_s",0) for r in res]
        rnks  = [r["retrieval_rank"] for r in nt4
                 if r.get("retrieval_rank") is not None and r.get("hit") == 1.0]
        mrr   = sum(1.0/x for x in rnks) / len(nt4) if rnks else None
        good  = sum(1 for r in nt4 if r.get("quality_verdict","").startswith("GOOD"))
        ok    = sum(1 for r in nt4 if r.get("quality_verdict","").startswith("OK"))
        bad   = sum(1 for r in nt4 if r.get("quality_verdict","").startswith("BAD"))
        return {
            "n":len(nt4),"n_total":len(res),
            "hit":  sum(hits)/len(hits)   if hits  else None,
            "cov":  sum(covs)/len(covs)   if covs  else None,
            "qual": sum(quals)/len(quals) if quals else None,
            "overall": sum(overs)/len(overs) if overs else None,
            "mrr":mrr,"latency":sum(lats)/len(lats) if lats else None,
            "good":good,"ok":ok,"bad":bad,
        }

    wb  = Workbook()
    ws1 = wb.active; ws1.title = "Dashboard"
    ws2 = wb.create_sheet("Results")
    ws3 = wb.create_sheet("Fact Check")

    # ── Dashboard ─────────────────────────────────────────────────────────
    ws1.sheet_view.showGridLines = False
    CW(ws1, [2,20,16,16,16,16,16,16,2])
    s   = _smry(results)
    row = 2

    ws1.merge_cells(f"B{row}:H{row}")
    ws1.row_dimensions[row].height = 44
    c = ws1[f"B{row}"]
    c.value = "  MMRAG  —  EVALUATION DASHBOARD"
    c.font  = FT(bold=True,color="white",size=22)
    c.fill  = F("navy"); c.alignment = AL(v="center")
    for col in range(2,9): ws1[f"{get_column_letter(col)}{row}"].fill = F("navy")
    row += 1

    ws1.merge_cells(f"B{row}:H{row}")
    ws1.row_dimensions[row].height = 20
    c = ws1[f"B{row}"]
    auto_count = sum(1 for r in results if r.get("quality_verdict","").endswith("*"))
    judge_note = f"Judge: llama3.1:8b (Ollama local)"
    if auto_count > 0:
        judge_note += f"  |  {auto_count} auto-scored (word-overlap, marked *)"
    c.value = (f"  Model: Llama-4-Scout (Groq)   —   {judge_note}   —   "
               f"Questions: {s['n_total']} total ({s['n']} scored)   —   Run: {datetime.date.today():%d %b %Y}")
    c.font  = FT(color="AECBF5",size=9.5,italic=True)
    c.fill  = F("navy2"); c.alignment = AL(v="center")
    for col in range(2,9): ws1[f"{get_column_letter(col)}{row}"].fill = F("navy2")
    row += 1

    kpis = [
        ("OVERALL\nSCORE",  s["overall"], "Tiers 1-3 average",         "steel", "lblue"),
        ("HIT RATE",        s["hit"],     "right slide found in top-K", "gfg",   "t1bg"),
        ("FACT COVERAGE",   s["cov"],     "key facts in the answer",    "afg",   "t3bg"),
        ("AVG QUALITY",     s["qual"],    "0=Bad  0.5=OK  1.0=Good",    "4527A0","F3E5F5"),
    ]
    ws1.row_dimensions[row].height = 18
    for i,(lbl,val,sub,fg,bg) in enumerate(kpis):
        cl = get_column_letter(2+i*2-1); ce = get_column_letter(2+i*2)
        ws1.merge_cells(f"{cl}{row}:{ce}{row}")
        c = ws1[f"{cl}{row}"]
        c.value = lbl; c.font = FT(bold=True,color="white",size=9)
        c.fill = F(fg); c.alignment = AL("center","center")
        ws1[f"{ce}{row}"].fill = F(fg)
    row += 1

    ws1.row_dimensions[row].height = 58
    for i,(lbl,val,sub,fg,bg) in enumerate(kpis):
        cl = get_column_letter(2+i*2-1); ce = get_column_letter(2+i*2)
        ws1.merge_cells(f"{cl}{row}:{ce}{row}")
        c = ws1[f"{cl}{row}"]
        c.value = Pct(val); c.font = FT(bold=True,color=fg,size=34)
        c.fill = F(bg); c.alignment = AL("center","center")
        ws1[f"{ce}{row}"].fill = F(bg)
    row += 2

    # Tier breakdown
    ws1.merge_cells(f"B{row}:H{row}")
    ws1.row_dimensions[row].height = 22
    c = ws1[f"B{row}"]
    c.value = "  TIER PERFORMANCE BREAKDOWN"
    c.font = FT(bold=True,color="white",size=11); c.fill = F("navy"); c.alignment = AL()
    for col in range(2,9): ws1[f"{get_column_letter(col)}{row}"].fill = F("navy")
    row += 1

    for col,hdr in zip(["B","C","D","E","F","G","H"],
                        ["Tier","Questions","Hit Rate","Fact Cov.","Avg Quality","Avg Score","Notes"]):
        c = ws1[f"{col}{row}"]
        c.value = hdr; c.font = FT(bold=True,color="white",size=9.5)
        c.fill = F("2C3E6B"); c.alignment = AL("center","center"); c.border = BD()
    row += 1

    tier_groups = defaultdict(list)
    for r in results: tier_groups[r.get("tier","unknown")].append(r)

    for tk,tm in TIER_META.items():
        tr = tier_groups.get(tk,[])
        if not tr: continue
        t_hits  = [r["hit"]           for r in tr if r.get("hit")           is not None]
        t_covs  = [r["fact_coverage"] for r in tr if r.get("fact_coverage") is not None]
        t_quals = [r["quality_score"] for r in tr if r.get("quality_score") is not None]
        t_overs = [r["overall"]       for r in tr if r.get("overall")       is not None]
        ws1.row_dimensions[row].height = 28
        vals = [tm["label"],str(len(tr)),
                Pct(sum(t_hits)/len(t_hits)   if t_hits  else None),
                Pct(sum(t_covs)/len(t_covs)   if t_covs  else None),
                Pct(sum(t_quals)/len(t_quals) if t_quals else None),
                Pct(sum(t_overs)/len(t_overs) if t_overs else None),
                tm["note"]]
        for col,val in zip(["B","C","D","E","F","G","H"],vals):
            c = ws1[f"{col}{row}"]
            c.value = val
            c.font = FT(bold=(col=="B"),color=tm["fg"] if col=="B" else "333333",size=9.5)
            c.fill = F(tm["bg"]); c.alignment = AL("left" if col in("B","H") else "center","center")
            c.border = BD()
        row += 1

    # ── Results sheet ─────────────────────────────────────────────────────
    ws2.sheet_view.showGridLines = False
    CW(ws2,[2,5,18,34,36,36,22,8,6,8,10,26])
    row = 2
    ws2.merge_cells(f"B{row}:M{row}")
    ws2.row_dimensions[row].height = 34
    c = ws2[f"B{row}"]
    c.value = "  DETAILED RESULTS — Every question scored"
    c.font = FT(bold=True,color="white",size=15); c.fill = F("navy"); c.alignment = AL(v="center")
    for col in range(2,14): ws2[f"{get_column_letter(col)}{row}"].fill = F("navy")
    row += 2

    hdrs = ["#","Tier","Question","Reference Answer","System Answer",
            "Top Retrieved Slide","Hit?","Rank","Fact Score","Quality","Overall","Missing Facts"]
    for i,hdr in enumerate(hdrs,2):
        c = ws2[f"{get_column_letter(i)}{row}"]
        c.value = hdr; c.font = FT(bold=True,color="white",size=9.5)
        c.fill = F("2C3E6B"); c.alignment = AL("center","center",wrap=True)
    ws2.freeze_panes = f"B{row+1}"
    row += 1

    for r in results:
        ws2.row_dimensions[row].height = 65
        tm  = TIER_META.get(r.get("tier",""),{"label":r.get("tier",""),"bg":"offwh","fg":"222222"})
        vrd = r.get("quality_verdict","")
        base_vrd = vrd.rstrip("*")
        v_bg = {"GOOD":"gbg","OK":"abg","BAD":"rbg"}.get(base_vrd,"offwh")
        v_fg = {"GOOD":"gfg","OK":"afg","BAD":"rfg"}.get(base_vrd,"555555")
        hit  = r.get("hit")
        h_bg = "gbg" if hit==1.0 else ("abg" if (hit is not None and 0.0 < hit < 1.0) else ("rbg" if hit==0.0 else "offwh"))
        h_fg = "gfg" if hit==1.0 else ("afg" if (hit is not None and 0.0 < hit < 1.0) else ("rfg" if hit==0.0 else "888888"))
        cov  = r.get("fact_coverage")
        c_bg = ("gbg" if cov and cov>=0.8 else ("abg" if cov and cov>=0.5 else "rbg")) if cov is not None else "offwh"
        ov   = r.get("overall")
        is_t4 = r.get("tier") == "4_stress"
        o_bg  = "graybg" if is_t4 else (("gbg" if ov and ov>=0.75 else ("abg" if ov and ov>=0.45 else "rbg")) if ov is not None else "offwh")

        def wc(col, val, cell_bg=tm["bg"], cell_fg="222222", bold=False, h="left"):
            c = ws2[f"{get_column_letter(col)}{row}"]
            c.value = val; c.fill = F(cell_bg)
            c.font = FT(bold=bold,color=cell_fg,size=9)
            c.alignment = AL(h,"center",wrap=True); c.border = BD()

        rank      = r.get("retrieval_rank")
        rank_str  = f"#{rank}" if rank else "--"
        retrieved = r.get("retrieved",[])
        top_slide = retrieved[0] if retrieved else "--"
        miss_str  = ", ".join(r.get("missing_facts",[])) or "All found"
        ov_label  = Pct(ov) + (" (excl)" if is_t4 else "")

        wc(2,  r["idx"],      bold=True, h="center")
        wc(3,  tm["label"],   bold=True, cell_fg=tm["fg"])
        wc(4,  r["question"], bold=True)
        wc(5,  r.get("reference",""))
        wc(6,  r.get("prediction",""))
        wc(7,  top_slide,     h="center")
        if hit == 1.0:
            hit_label = "Hit"
        elif hit == 0.0:
            hit_label = "Miss"
        elif hit is not None:
            hit_label = f"Recall\n{hit:.0%}"  # partial recall for Tier 3
        else:
            hit_label = "--"
        wc(8,  hit_label, cell_bg=h_bg,cell_fg=h_fg,bold=True,h="center")
        wc(9,  rank_str,      h="center")
        wc(10, Pct(cov),      cell_bg=c_bg, h="center")
        wc(11, vrd,           cell_bg=v_bg,cell_fg=v_fg,bold=True,h="center")
        wc(12, ov_label,      cell_bg=o_bg, h="center")
        wc(13, miss_str,      cell_fg=("rfg" if r.get("missing_facts") else "gfg"))
        row += 1

    # ── Fact Check sheet ──────────────────────────────────────────────────
    ws3.sheet_view.showGridLines = False
    CW(ws3,[2,5,18,44,32,14])
    row = 2
    ws3.merge_cells(f"B{row}:G{row}")
    ws3.row_dimensions[row].height = 34
    c = ws3[f"B{row}"]
    c.value = "  FACT CHECK — Key facts found or missing in each answer"
    c.font = FT(bold=True,color="white",size=15); c.fill = F("navy"); c.alignment = AL(v="center")
    for col in range(2,8): ws3[f"{get_column_letter(col)}{row}"].fill = F("navy")
    row += 2

    for i,hdr in enumerate(["#","Tier","Question","Result","Score"],2):
        c = ws3[f"{get_column_letter(i)}{row}"]
        c.value = hdr; c.font = FT(bold=True,color="white",size=9.5)
        c.fill = F("2C3E6B"); c.alignment = AL("center","center")
    ws3.freeze_panes = f"B{row+1}"
    row += 1

    for r in results:
        cov = r.get("fact_coverage")
        if cov is None: continue
        tm   = TIER_META.get(r.get("tier",""),{"label":r.get("tier",""),"bg":"offwh","fg":"222222"})
        miss = r.get("missing_facts",[])
        c_bg = "gbg" if cov>=0.8 else ("abg" if cov>=0.5 else "rbg")
        c_fg = "gfg" if cov>=0.8 else ("afg" if cov>=0.5 else "rfg")
        ws3.row_dimensions[row].height = 36

        def wc(col,val,cell_bg=tm["bg"],cell_fg="222222",bold=False):
            c = ws3[f"{get_column_letter(col)}{row}"]
            c.value = val; c.fill = F(cell_bg)
            c.font = FT(bold=bold,color=cell_fg,size=9.5)
            c.alignment = AL("center" if col<=3 else "left","center",wrap=True); c.border = BD()

        result_txt = (", ".join(f'Missing: "{m}"' for m in miss) if miss else "All key facts found")
        wc(2,r["idx"],bold=True)
        wc(3,tm["label"],bold=True,cell_fg=tm["fg"])
        wc(4,r["question"])
        wc(5,result_txt,cell_fg=("rfg" if miss else "gfg"))
        wc(6,Pct(cov),cell_bg=c_bg,cell_fg=c_fg,bold=True)
        row += 1

    wb.save(output_path)
    print(f"\n✅ Excel saved to: {output_path}")

# ─────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("  MM-RAG LOCAL EVALUATION")
    print("=" * 60)
    print()
    print("Loading backend (CLIP + MiniLM + EasyOCR)...")
    backend = MMRAGBackend()

    slides = backend.list_slides()
    if not slides:
        print("❌ No slides indexed. Add slides to mmrag_store/pages/ first.")
        sys.exit(1)
    print(f"✅ {len(slides)} slides indexed")
    print(f"✅ {len(golden_qa)} questions loaded")
    print()

    results = run_evaluation(backend)

    # Autosave results so Excel can be re-exported without re-running
    with open("mmrag_results.json", "w") as f:
        json.dump(results, f, indent=2)
    print("✅ Results saved to mmrag_results.json")

    export_excel(results, "mmrag_evaluation.xlsx")
